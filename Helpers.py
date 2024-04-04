import time 
from datetime import datetime as dt 
import openpyxl

import uuid 

#for sending in eamils 
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# Get today's date
today = dt.now()
today_day_month = today.strftime('%d-%m') 

#openpyxl code
path=r"C:\Users\Tanmay\Desktop\Personal\Network\Nework.xlsx"  #raw string to prevent escaping of characters
wb=openpyxl.load_workbook(path)
sheet=wb.active
max_row=sheet.max_row

import time

def ticktock(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        print(f"{func.__name__} took {end_time - start_time:.6f} seconds to execute")
        return result
    return wrapper

#**Helper Functions** 

#check_birthdays from openpyxl  
@ticktock
def check_birthdays(Wb,Sheet,Max_row):  
    sheet = Sheet 
    max_row = Max_row
    wb=Wb 

    list_of_indices = []
    for i in range(1, max_row+1):
        cell_value = sheet.cell(row=i, column=4).value  # Assuming the date is in the 4th column (column index 3)
        if cell_value and isinstance(cell_value, dt):  
            #print("Yes") # Check if the cell contains a datetime object
            cell_value_str = cell_value.strftime('%d-%m-%Y')  # Convert datetime object to string
            cell_day_month = dt.strptime(cell_value_str, '%d-%m-%Y').strftime('%d-%m') 
            if cell_day_month == today_day_month:
        
                list_of_indices.append(i) 
           
    return list_of_indices  

#Function to check birthdays in spreadsheet 

@ticktock
def check_birthdays_sheets(birthday_values, today_day_month):
    list_of_indices = [] 
    print("Checking for Birthdays today.....")
    
    for i, cell in enumerate(birthday_values):
        if cell:
            cell_value = str(cell[0]).strip()
            
            try:
                cell_value_dt = dt.strptime(cell_value, '%m/%d/%Y')
                cell_day_month = cell_value_dt.strftime('%d-%m')
                
                if cell_day_month == today_day_month:
                    list_of_indices.append(i + 2)
                    print(f"Row numbers to append: {i + 2}")
            except ValueError:
                print(f"Error parsing date in row {i + 1}: {cell_value}")

    return list_of_indices
#Function to calculate days left in deadline  


def days_left_in_deadline(deadline,today):
    if deadline:
        time_diff=(deadline-today)
    else:
        time_diff = None 
    #Converting the time duration to +ve duration string  
    overdue = False
    if time_diff: 
        if time_diff.days < 0:
            overdue=True
            time_diff = f"Overdue by {abs(time_diff)})" 
        else:
            time_diff = f"(Due in {abs(time_diff)})" 
    else:
        f"No Deadline Yet"
    return time_diff,overdue 


#helper functions to add tasks,events and deadlines 
def add_task(tasks,task_name, task_summary, start_hour, end_hour):
    
    task_id = str(uuid.uuid4())
    tasks.append({  # Append the new task dictionary to the list
        'id': task_id,
        'task': task_name,
        'summary': task_summary,
        'start_hour': start_hour,
        'end_hour': end_hour,
        'done': False
    })


def add_event(events,event_name, task_summary, start_date, end_date):
    
    event_id = str(uuid.uuid4())
    events.append({
        'id': event_id,
        'task': event_name,
        'summary': task_summary,
        'start_date': start_date,
        'end_date': end_date,
        'done': False
    })
#getting the time remaining till for the day to end
def get_time_remaining():
    #get current time 
    now=dt.now()
    EOD=dt.combine(now.date(),dt.max.time()) 

    time_remaining=EOD-now 
    return str(time_remaining)[:8].rstrip('.') 

#Sending in email notifications 

def send_email( subject, body, to_email):
    # Email configuration
    smtp_server = 'smtp.gmail.com'  # Update with your SMTP server
    smtp_port = 587  # Update with your SMTP port
    sender_email = 'onmailharshit@gmail.com'  # Update with your email address
    sender_password = 'sasw syyp dtnz xlmd'  # Update with your email password

    # Create a multipart message and set headers
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = to_email
    msg['Subject'] = subject

    # Add body to email
    msg.attach(MIMEText(body, 'plain'))

    # Create SMTP session for sending the email
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_email, msg.as_string())
        server.quit()
        print('Email sent successfully!')
    except Exception as e:
        print(f'Error sending email: {e}')

#timing functrion 

